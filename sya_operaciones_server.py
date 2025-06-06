# sya_operaciones_server.py
import os
import logging
from datetime import datetime
import pandas as pd
import openpyxl
from flask import Flask, request, jsonify, send_file, send_from_directory
import zipfile

app = Flask(__name__)

# Usar rutas absolutas
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(BASE_DIR, "registros_trabajo.xlsx")
REQUERIMIENTOS_EXCEL_FILE = os.path.join(BASE_DIR, "requerimientos_obra.xlsx")
MATERIALES_CSV_PATH = os.path.join(BASE_DIR, "operaciones_materiales.csv")
EQUIPOS_CSV_PATH = os.path.join(BASE_DIR, "operaciones_equipos.csv")
VEHICULOS_CSV_PATH = os.path.join(BASE_DIR, "operaciones_vehiculos.csv")
PERSONAL_CSV_PATH = os.path.join(BASE_DIR, "operaciones_personal.csv")

# Archivo Excel y directorio para registros de choferes
REGISTROS_CHOFERES_EXCEL = os.path.join(BASE_DIR, "registros_choferes.xlsx")
FOTOS_VEHICULOS_DIR = os.path.join(BASE_DIR, "fotos_vehiculos")

# Ruta al archivo CSV de conductores y vehiculos
CONDUCTORES_CSV_PATH = os.path.join(BASE_DIR, "aem_conductores.csv")
VEHICULOS_INFO_CSV_PATH = os.path.join(BASE_DIR, "aem_vehiculos.csv")

# Configuración de logging
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s')

def inicializar_excel():
    """Inicializa los archivos Excel si no existen."""
    # Inicializar Excel de Reporte Diario
    if not os.path.exists(EXCEL_FILE):
        logging.info(f"Creando archivo Excel de reporte diario en: {EXCEL_FILE}")
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Reporte Principal"
        headers_reporte = [
            "Fecha", "Código Obra", "Nombre Ingeniero",
            "Nombre Supervisor", "Actividad Principal",
            "Supervisor Presente", "Avance Diario",
            "Incidentes", "Plan Siguiente Día", "Observaciones"
        ]
        ws1.append(headers_reporte)
        ws2 = wb.create_sheet(title="Materiales Usados")
        headers_materiales = ["Fecha", "Código Obra", "Nombre Ingeniero"]
        ws2.append(headers_materiales)
        ws3 = wb.create_sheet(title="Equipos Usados")
        headers_equipos = ["Fecha", "Código Obra", "Nombre Ingeniero"]
        ws3.append(headers_equipos)
        ws4 = wb.create_sheet(title="Vehículos Usados")
        headers_vehiculos = ["Fecha", "Código Obra", "Nombre Ingeniero"]
        ws4.append(headers_vehiculos)
        ws5 = wb.create_sheet(title="Personal de Campo")
        headers_personal = ["Fecha", "Código Obra", "Nombre Ingeniero"]
        ws5.append(headers_personal)
        wb.save(EXCEL_FILE)
    else:
        logging.info(f"El archivo Excel de reporte diario ya existe en: {EXCEL_FILE}")

    # Inicializar Excel de Requerimientos
    if not os.path.exists(REQUERIMIENTOS_EXCEL_FILE):
        logging.info(f"Creando archivo Excel de requerimientos en: {REQUERIMIENTOS_EXCEL_FILE}")
        wb_req = openpyxl.Workbook()
        ws_req = wb_req.active
        ws_req.title = "Requerimientos"
        headers_requerimientos_inicial = ["Fecha", "Código Obra", "Nombre Ingeniero"]
        ws_req.append(headers_requerimientos_inicial)
        wb_req.save(REQUERIMIENTOS_EXCEL_FILE)
    else:
        logging.info(f"El archivo Excel de requerimientos ya existe en: {REQUERIMIENTOS_EXCEL_FILE}")

    # Inicializar Excel de Registros de Choferes
# Inicializar Excel de Registros de Choferes
    if not os.path.exists(REGISTROS_CHOFERES_EXCEL):
        logging.info(f"Creando archivo Excel de registros de choferes en: {REGISTROS_CHOFERES_EXCEL}")
        wb_choferes = openpyxl.Workbook()
        ws_choferes = wb_choferes.active
        ws_choferes.title = "Registros"
        headers_choferes = [
            "Fecha", "Nombre del Chofer", "Vehículo", "Placa", "Fecha de Salida", 
            "Hora de Salida", "Ubicación Inicial", "Kilometraje Inicial", 
            "Observaciones Salida", "Fecha de Llegada", "Hora de Retorno",
            "Ubicación Final", "Kilometraje Final", "Observaciones Llegada"
        ]
        ws_choferes.append(headers_choferes)
        wb_choferes.save(REGISTROS_CHOFERES_EXCEL)
    else:
        logging.info(f"El archivo Excel de registros de choferes ya existe en: {REGISTROS_CHOFERES_EXCEL}")

    # Crear directorio de fotos si no existe
    if not os.path.exists(FOTOS_VEHICULOS_DIR):
        os.makedirs(FOTOS_VEHICULOS_DIR)
        logging.info(f"Directorio de fotos creado: {FOTOS_VEHICULOS_DIR}")

def actualizar_cabeceras_materiales(ws, num_materiales):
    """Actualiza las cabeceras de la hoja de materiales."""
    headers = list(ws.rows)[0]
    num_headers_actuales = len(headers)

    ultimo_material = 0
    for header in headers:
        header_value = header.value
        if header_value and header_value.startswith("Material"):
            try:
                numero = int(header_value.split(" ")[1])
                ultimo_material = max(ultimo_material, numero)
            except ValueError:
                pass

    if ultimo_material >= num_materiales:
        return

    for i in range(ultimo_material + 1, num_materiales + 1):
        ws.cell(row=1, column=num_headers_actuales + 1, value=f"Material {i}")
        ws.cell(row=1, column=num_headers_actuales + 2, value=f"Unidad {i}")
        ws.cell(row=1, column=num_headers_actuales + 3, value=f"Cantidad {i}")
        num_headers_actuales += 3

def actualizar_cabeceras_equipos(ws, num_equipos):
    """Actualiza las cabeceras de la hoja de equipos."""
    headers = list(ws.rows)[0]
    num_headers_actuales = len(headers)

    ultimo_equipo = 0
    for header in headers:
        header_value = header.value
        if header_value and header_value.startswith("Equipo"):
            try:
                numero = int(header_value.split(" ")[1])
                ultimo_equipo = max(ultimo_equipo, numero)
            except ValueError:
                pass

    if ultimo_equipo >= num_equipos:
        return

    for i in range(ultimo_equipo + 1, num_equipos + 1):
        ws.cell(row=1, column=num_headers_actuales + 1, value=f"Equipo {i}")
        ws.cell(row=1, column=num_headers_actuales + 2, value=f"Cantidad {i}")
        ws.cell(row=1, column=num_headers_actuales + 3, value=f"Propiedad {i}")
        num_headers_actuales += 3

def actualizar_cabeceras_vehiculos(ws, num_vehiculos):
    """Actualiza las cabeceras de la hoja de vehículos."""
    headers = list(ws.rows)[0]
    num_headers_actuales = len(headers)

    ultimo_vehiculo = 0
    for header in headers:
        header_value = header.value
        if header_value and header_value.startswith("Vehículo"):
            try:
                numero = int(header_value.split(" ")[1])
                ultimo_vehiculo = max(ultimo_vehiculo, numero)
            except ValueError:
                pass

    if ultimo_vehiculo >= num_vehiculos:
        return

    for i in range(ultimo_vehiculo + 1, num_vehiculos + 1):
        ws.cell(row=1, column=num_headers_actuales + 1, value=f"Vehículo {i}")
        ws.cell(row=1, column=num_headers_actuales + 2, value=f"Placa {i}")
        ws.cell(row=1, column=num_headers_actuales + 3, value=f"Propiedad {i}")
        num_headers_actuales += 3

def actualizar_cabeceras_personal(ws, num_personal):
    """Actualiza las cabeceras de la hoja de personal."""
    headers = list(ws.rows)[0]
    num_headers_actuales = len(headers)

    ultimo_personal = 0
    for header in headers:
        header_value = header.value
        if header_value and header_value.startswith("Personal"):
            try:
                numero = int(header_value.split(" ")[1])
                ultimo_personal = max(ultimo_personal, numero)
            except ValueError:
                pass

    if ultimo_personal >= num_personal:
        return

    for i in range(ultimo_personal + 1, num_personal + 1):
        ws.cell(row=1, column=num_headers_actuales + 1, value=f"Personal {i}")
        ws.cell(row=1, column=num_headers_actuales + 2, value=f"Categoría {i}")
        ws.cell(row=1, column=num_headers_actuales + 3, value=f"Horas extras {i}")
        num_headers_actuales += 3

def actualizar_cabeceras_requerimientos(ws, num_items):
    """Actualiza las cabeceras de la hoja de requerimientos."""
    headers = list(ws.rows)[0]
    num_headers_actuales = len(headers)

    ultimo_item = 0
    for header in headers:
        header_value = header.value
        if header_value and header_value.startswith("Artículo"):
            try:
                numero = int(header_value.split(" ")[1])
                ultimo_item = max(ultimo_item, numero)
            except ValueError:
                pass

    if ultimo_item >= num_items:
        return

    for i in range(ultimo_item + 1, num_items + 1):
        ws.cell(row=1, column=num_headers_actuales + 1, value=f"Artículo {i}")
        ws.cell(row=1, column=num_headers_actuales + 2, value=f"Unidad {i}")
        ws.cell(row=1, column=num_headers_actuales + 3, value=f"Cantidad {i}")
        num_headers_actuales += 3


def procesar_datos(datos):
    """Procesa los datos del reporte diario."""
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws_reporte = wb["Reporte Principal"]
        ws_materiales = wb["Materiales Usados"]
        ws_equipos = wb["Equipos Usados"]
        ws_vehiculos = wb["Vehículos Usados"]
        ws_personal = wb["Personal de Campo"]

        # Preparar fila de datos para "Reporte Principal"
        fila_reporte = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', ''),
            datos.get('nombre_supervisor', ''),
            datos.get('actividad_principal', ''),
            'Sí' if datos.get('supervisor_presente', False) else 'No',
            datos.get('avance_diario', ''),
            datos.get('incidentes', ''),
            datos.get('siguiente_dia', ''),
            datos.get('observaciones', '')
        ]
        ws_reporte.append(fila_reporte)

        # Procesar materiales usados
        materiales = datos.get('materiales_usados', [])
        actualizar_cabeceras_materiales(ws_materiales, len(materiales))
        fila_materiales = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', '')
        ]
        for material in materiales:
            fila_materiales.extend([material['nombre'], material['unidad'], material['cantidad']])
        ws_materiales.append(fila_materiales)

        # Procesar equipos usados
        equipos = datos.get('equipos_usados', [])
        actualizar_cabeceras_equipos(ws_equipos, len(equipos))
        fila_equipos = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', '')
        ]
        for equipo in equipos:
            fila_equipos.extend([equipo['nombre'], equipo['cantidad'], equipo['propiedad']])
        ws_equipos.append(fila_equipos)

        # Procesar vehículos usados
        vehiculos = datos.get('vehiculos_usados', [])
        actualizar_cabeceras_vehiculos(ws_vehiculos, len(vehiculos))
        fila_vehiculos = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', '')
        ]
        for vehiculo in vehiculos:
            fila_vehiculos.extend([vehiculo['nombre'], vehiculo['placa'], vehiculo['propiedad']])
        ws_vehiculos.append(fila_vehiculos)

        # Procesar personal de campo
        personal_campo = datos.get('personal_de_campo', [])
        actualizar_cabeceras_personal(ws_personal, len(personal_campo))
        fila_personal = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', '')
        ]
        for personal in personal_campo:
            fila_personal.extend([personal['nombre_completo'], personal['categoria'], personal['horas_extras']])
        ws_personal.append(fila_personal)

        wb.save(EXCEL_FILE)
        logging.info(f"Datos recibidos de {datos.get('nombre_ingeniero', 'Unknown')} procesados exitosamente")

    except Exception as e:
        logging.error(f"Error al procesar datos: {str(e)}")

def procesar_requerimientos(datos):
    """Procesa los datos de requerimientos."""
    logging.info("Datos de requerimientos recibidos:")
    logging.info(datos)
    try:
        wb_req = openpyxl.load_workbook(REQUERIMIENTOS_EXCEL_FILE)
        ws_requerimientos = wb_req["Requerimientos"]

        requerimientos = datos.get('requerimientos', [])
        actualizar_cabeceras_requerimientos(ws_requerimientos, len(requerimientos))

        fila_requerimientos = [
            datetime.strptime(datos.get('fecha', ''), '%d/%m/%Y').date(),
            datos.get('codigo_obra', ''),
            datos.get('nombre_ingeniero', '')
        ]
        for req in requerimientos:
            fila_requerimientos.extend([req['nombre'], req['unidad'], req['cantidad']])
        ws_requerimientos.append(fila_requerimientos)

        wb_req.save(REQUERIMIENTOS_EXCEL_FILE)
        logging.info(f"Requerimientos recibidos de {datos.get('nombre_ingeniero', 'Unknown')} procesados exitosamente")

    except Exception as e:
        logging.exception(f"Error al procesar requerimientos: {str(e)}")

def descargar_excel_flask():
    """Descarga el archivo Excel principal."""
    try:
        logging.info(f"Intentando enviar archivo: {EXCEL_FILE}")
        return send_file(EXCEL_FILE, as_attachment=True)
    except Exception as e:
        logging.error(f"Error al generar descarga de Excel: {str(e)}")
        return str(e), 500

def descargar_requerimientos_excel_flask():
    """Descarga el archivo Excel de requerimientos."""
    try:
        logging.info(f"Intentando enviar archivo de requerimientos: {REQUERIMIENTOS_EXCEL_FILE}")
        return send_file(REQUERIMIENTOS_EXCEL_FILE, as_attachment=True, download_name='requerimientos_obra.xlsx')
    except Exception as e:
        logging.error(f"Error al generar descarga de Excel de requerimientos: {str(e)}")
        return str(e), 500


def agregar_nuevo_material_csv(nombre_material, unidad):
    """Agrega un nuevo material al archivo CSV."""
    try:
        df = pd.read_csv(MATERIALES_CSV_PATH)
        nuevo_material = pd.DataFrame([{'nombre_material': nombre_material, 'unidad': unidad}])
        df = pd.concat([df, nuevo_material], ignore_index=True)
        df.to_csv(MATERIALES_CSV_PATH, index=False)
        logging.info(f"Nuevo material '{nombre_material}' agregado a {MATERIALES_CSV_PATH}")
        return True
    except Exception as e:
        logging.error(f"Error al agregar nuevo material a CSV: {str(e)}")
        return False

def agregar_nuevo_equipo_csv(nombre_equipo, propiedad):
    """Agrega un nuevo equipo al archivo CSV."""
    try:
        df = pd.read_csv(EQUIPOS_CSV_PATH)
        nuevo_equipo = pd.DataFrame([{'nombre_equipo': nombre_equipo, 'propiedad': propiedad}])
        df = pd.concat([df, nuevo_equipo], ignore_index=True)
        df.to_csv(EQUIPOS_CSV_PATH, index=False)
        logging.info(f"Nuevo equipo '{nombre_equipo}' agregado a {EQUIPOS_CSV_PATH}")
        return True
    except Exception as e:
        logging.error(f"Error al agregar nuevo equipo a CSV: {str(e)}")
        return False

def agregar_nuevo_vehiculo_csv(nombre_vehiculo, placa, propiedad):
    """Agrega un nuevo vehículo al archivo CSV."""
    try:
        df = pd.read_csv(VEHICULOS_CSV_PATH)
        nuevo_vehiculo = pd.DataFrame([{'nombre_vehiculo': nombre_vehiculo, 'placa': placa, 'propiedad': propiedad}])
        df = pd.concat([df, nuevo_vehiculo], ignore_index=True)
        df.to_csv(VEHICULOS_CSV_PATH, index=False)
        logging.info(f"Nuevo vehículo '{nombre_vehiculo}' agregado a {VEHICULOS_CSV_PATH}")
        return True
    except Exception as e:
        logging.error(f"Error al agregar nuevo vehículo a CSV: {str(e)}")
        return False

def agregar_nuevo_personal_csv(apellido_paterno, apellido_materno, nombres, categoria):
    """Agrega un nuevo personal al archivo CSV."""
    try:
        df = pd.read_csv(PERSONAL_CSV_PATH)
        nuevo_personal = pd.DataFrame([{
            'AP. PATERNO': apellido_paterno,
            'AP. MATERNO': apellido_materno,
            'NOMBRES': nombres,
            'CATEGORIA': categoria
        }])
        df = pd.concat([df, nuevo_personal], ignore_index=True)
        df.to_csv(PERSONAL_CSV_PATH, index=False)
        logging.info(f"Nuevo personal '{nombres} {apellido_paterno}' agregado a {PERSONAL_CSV_PATH}")
        return True
    except Exception as e:
        logging.error(f"Error al agregar nuevo personal a CSV: {str(e)}")
        return False


# Inicializar Excel al inicio
inicializar_excel()

# Rutas de la API
@app.route('/api/materiales', methods=['GET'])
def get_materiales():
    """Obtiene la lista de materiales."""
    try:
        if not os.path.exists(MATERIALES_CSV_PATH):
            return jsonify({"error": f"No se encontró el archivo de materiales en {MATERIALES_CSV_PATH}"}), 404
        df = pd.read_csv(MATERIALES_CSV_PATH)
        materiales = df.to_dict(orient='records')
        return jsonify(materiales)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/equipos', methods=['GET'])
def get_equipos():
    """Obtiene la lista de equipos."""
    try:
        if not os.path.exists(EQUIPOS_CSV_PATH):
             return jsonify({"error": f"No se encontró el archivo de equipos en {EQUIPOS_CSV_PATH}"}), 404
        df = pd.read_csv(EQUIPOS_CSV_PATH)
        equipos = df.to_dict(orient='records')
        return jsonify(equipos)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/vehiculos', methods=['GET'])
def get_vehiculos():
    """Obtiene la lista de vehículos."""
    try:
        if not os.path.exists(VEHICULOS_CSV_PATH):
            return jsonify({"error": f"No se encontró el archivo de vehículos en {VEHICULOS_CSV_PATH}"}), 404
        df = pd.read_csv(VEHICULOS_CSV_PATH)
        vehiculos = df.to_dict(orient='records')
        return jsonify(vehiculos)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/personal', methods=['GET'])
def get_personal():
    """Obtiene la lista de personal."""
    try:
        if not os.path.exists(PERSONAL_CSV_PATH):
            return jsonify({"error": f"No se encontró el archivo de personal en {PERSONAL_CSV_PATH}"}), 404
        df = pd.read_csv(PERSONAL_CSV_PATH)
        personal = df.to_dict(orient='records')
        return jsonify(personal)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/materiales/new', methods=['POST'])
def new_material():
    """Agrega un nuevo material."""
    data = request.json
    if not data or 'nombre_material' not in data or 'unidad' not in data:
        return jsonify({"error": "Nombre de material y unidad son requeridos"}), 400
    if agregar_nuevo_material_csv(data['nombre_material'], data['unidad']):
        return jsonify({"status": "success"}), 201
    else:
        return jsonify({"error": "Error al agregar nuevo material"}), 500

@app.route('/api/equipos/new', methods=['POST'])
def new_equipo():
    """Agrega un nuevo equipo."""
    data = request.json
    if not data or 'nombre_equipo' not in data or 'propiedad' not in data:
        return jsonify({"error": "Nombre de equipo y propiedad son requeridos"}), 400
    if agregar_nuevo_equipo_csv(data['nombre_equipo'], data['propiedad']):
        return jsonify({"status": "success"}), 201
    else:
        return jsonify({"error": "Error al agregar nuevo equipo"}), 500

@app.route('/api/vehiculos/new', methods=['POST'])
def new_vehiculo():
    """Agrega un nuevo vehículo."""
    data = request.json
    if not data or 'nombre_vehiculo' not in data or 'placa' not in data or 'propiedad' not in data:
        return jsonify({"error": "Nombre de vehículo, placa y propiedad son requeridos"}), 400
    if agregar_nuevo_vehiculo_csv(data['nombre_vehiculo'], data['placa'], data['propiedad']):
        return jsonify({"status": "success"}), 201
    else:
        return jsonify({"error": "Error al agregar nuevo vehículo"}), 500

@app.route('/api/personal/new', methods=['POST'])
def new_personal():
    """Agrega nuevo personal."""
    data = request.json
    if not data or 'nombre_completo' not in data or 'categoria' not in data:
        return jsonify({"error": "Nombre completo y categoría de personal son requeridos"}), 400

    nombre_completo = data['nombre_completo']
    partes_nombre = nombre_completo.split(',')
    if len(partes_nombre) != 2:
        return jsonify({"error": "Formato de nombre incorrecto. Debe ser 'apellido_paterno apellido_materno, nombres'"}), 400

    apellidos_str, nombres = partes_nombre
    apellidos_partes = apellidos_str.strip().split()
    if not apellidos_partes:
        return jsonify({"error": "Apellido paterno requerido"}), 400
    apellido_paterno = apellidos_partes[0]
    apellido_materno = apellidos_partes[1] if len(apellidos_partes) > 1 else ''

    if agregar_nuevo_personal_csv(apellido_paterno, apellido_materno, nombres.strip(), data['categoria']):
        return jsonify({"status": "success"}), 201
    else:
        return jsonify({"error": "Error al agregar nuevo personal"}), 500


@app.route('/recibir-datos', methods=['POST'])
def recibir_datos():
    """Recibe los datos del reporte diario."""
    datos = request.json
    procesar_datos(datos)
    return jsonify({"status": "success"})

@app.route('/recibir-requerimientos', methods=['POST'])
def recibir_requerimientos_route():
    """Recibe los datos de requerimientos."""
    datos = request.json
    print("Datos recibidos en /recibir-requerimientos:", datos)
    procesar_requerimientos(datos)
    return jsonify({"status": "success"})

@app.route('/descargar-excel', methods=['GET'])
def descargar_excel_route():
    """Descarga el archivo Excel principal."""
    return descargar_excel_flask()

@app.route('/descargar-requerimientos-excel', methods=['GET'])
def descargar_requerimientos_excel_route():
    """Descarga el archivo Excel de requerimientos."""
    return descargar_requerimientos_excel_flask()

# Funciones y rutas para la app de choferes
def procesar_datos_choferes(data, files):
    """Procesa los datos del formulario de choferes o solo guarda fotos si se proporciona un row_idx."""
    try:
        wb_choferes = openpyxl.load_workbook(REGISTROS_CHOFERES_EXCEL)
        ws_choferes = wb_choferes.active

        nombre_chofer = data.get("nombre_chofer")
        placa = data.get("placa")
        tipo_formulario = data.get("tipo_formulario")
        row_idx = data.get("row_idx")  # Identificador de fila (opcional)

        # Función para generar el nombre de la subcarpeta
        def generar_nombre_subcarpeta(fecha_salida, nombre_chofer, placa):
            fecha_salida_filename = fecha_salida.replace("-", "")
            nombre_chofer_filename = nombre_chofer.lower().replace(" ", "-")
            placa_filename = placa.replace(" ", "-")
            return f"{fecha_salida_filename}_{nombre_chofer_filename}_{placa_filename}"

        # Si solo se envía row_idx y fotos (segunda solicitud para llegada)
        if row_idx and not tipo_formulario:
            try:
                row_idx = int(row_idx)
                if row_idx <= 1 or row_idx > ws_choferes.max_row:
                    return False, "Índice de fila inválido."

                # Obtener la fecha de salida desde el Excel (columna 5: Fecha de Salida)
                fecha_salida_excel = ws_choferes.cell(row=row_idx, column=5).value
                if isinstance(fecha_salida_excel, datetime):
                    fecha_salida_str = fecha_salida_excel.strftime("%Y-%m-%d")
                else:
                    fecha_salida_str = str(fecha_salida_excel)

                # Generar el nombre de la subcarpeta
                subcarpeta_nombre = generar_nombre_subcarpeta(fecha_salida_str, nombre_chofer, placa)
                subcarpeta_path = os.path.join(FOTOS_VEHICULOS_DIR, subcarpeta_nombre)

                # Crear la subcarpeta si no existe
                if not os.path.exists(subcarpeta_path):
                    os.makedirs(subcarpeta_path)
                    logging.info(f"Subcarpeta creada: {subcarpeta_path}")

                # Guardar las fotos de llegada en la subcarpeta
                for i in range(1, 5):
                    foto_key = f"foto_km_final_{i}"
                    foto_fin = files.get(foto_key)
                    if foto_fin:
                        original_extension = os.path.splitext(foto_fin.filename)[1] if foto_fin.filename else ".jpg"
                        filename_fin = f"{subcarpeta_nombre}_llegada_{i}{original_extension}"
                        path_fin = os.path.join(subcarpeta_path, filename_fin)
                        foto_fin.save(path_fin)
                        logging.info(f"Foto de fin {i} guardada en {path_fin} para fila {row_idx}")
                return True, "Fotos de llegada guardadas correctamente."
            except ValueError:
                return False, "Índice de fila debe ser un número entero."

        # Lógica para formulario de salida
        if tipo_formulario == "salida":
            fecha_salida = data.get("fecha_salida")
            # Generar el nombre de la subcarpeta
            subcarpeta_nombre = generar_nombre_subcarpeta(fecha_salida, nombre_chofer, placa)
            subcarpeta_path = os.path.join(FOTOS_VEHICULOS_DIR, subcarpeta_nombre)

            # Crear la subcarpeta si no existe
            if not os.path.exists(subcarpeta_path):
                os.makedirs(subcarpeta_path)
                logging.info(f"Subcarpeta creada: {subcarpeta_path}")

            # Guardar las fotos de salida en la subcarpeta
            for i in range(1, 5):
                foto_key = f"foto_km_inicial_{i}"
                foto_inicio = files.get(foto_key)
                if foto_inicio:
                    original_extension = os.path.splitext(foto_inicio.filename)[1] if foto_inicio.filename else ".jpg"
                    filename_inicio = f"{subcarpeta_nombre}_salida_{i}{original_extension}"
                    path_inicio = os.path.join(subcarpeta_path, filename_inicio)
                    foto_inicio.save(path_inicio)
                    logging.info(f"Foto de inicio {i} guardada en {path_inicio}")

            # Guardar los datos en el Excel
            fecha_salida_date = datetime.strptime(fecha_salida, "%Y-%m-%d").date()
            fila_salida = [
                fecha_salida_date,  # Fecha
                nombre_chofer,
                data.get("vehiculo"),
                placa,
                fecha_salida_date,  # Fecha de Salida
                data.get("hora_salida"),
                data.get("ubicacion_inicial"),
                data.get("km_inicial"),
                data.get("observaciones_salida"),
                None, None, None, None, None
            ]
            ws_choferes.append(fila_salida)
            wb_choferes.save(REGISTROS_CHOFERES_EXCEL)
            logging.info(f"Datos de salida guardados en nueva fila.")
            return True, "Datos de salida guardados correctamente."

        # Lógica para formulario de llegada
        elif tipo_formulario == "llegada":
            ultimo_registro = None
            for row_idx in range(ws_choferes.max_row, 1, -1):
                if (ws_choferes.cell(row=row_idx, column=2).value == nombre_chofer and
                    ws_choferes.cell(row=row_idx, column=4).value == placa):
                    ultimo_registro = row_idx
                    break

            if ultimo_registro:
                if (ws_choferes.cell(row=ultimo_registro, column=10).value is None and
                    ws_choferes.cell(row=ultimo_registro, column=11).value is None and
                    ws_choferes.cell(row=ultimo_registro, column=12).value is None and
                    ws_choferes.cell(row=ultimo_registro, column=13).value is None and
                    ws_choferes.cell(row=ultimo_registro, column=14).value is None):
                    # Actualizar los datos de llegada en el Excel
                    ws_choferes.cell(row=ultimo_registro, column=10).value = datetime.strptime(data.get("fecha_llegada"), "%Y-%m-%d").date()
                    ws_choferes.cell(row=ultimo_registro, column=11).value = data.get("hora_retorno")
                    ws_choferes.cell(row=ultimo_registro, column=12).value = data.get("ubicacion_final")
                    ws_choferes.cell(row=ultimo_registro, column=13).value = data.get("km_final")
                    ws_choferes.cell(row=ultimo_registro, column=14).value = data.get("observaciones_llegada")
                    wb_choferes.save(REGISTROS_CHOFERES_EXCEL)
                    logging.info(f"Datos de llegada actualizados en fila {ultimo_registro}.")

                    # Obtener la fecha de salida desde el Excel (columna 5: Fecha de Salida)
                    fecha_salida_excel = ws_choferes.cell(row=ultimo_registro, column=5).value
                    if isinstance(fecha_salida_excel, datetime):
                        fecha_salida_str = fecha_salida_excel.strftime("%Y-%m-%d")
                    else:
                        fecha_salida_str = str(fecha_salida_excel)

                    # Generar el nombre de la subcarpeta
                    subcarpeta_nombre = generar_nombre_subcarpeta(fecha_salida_str, nombre_chofer, placa)
                    subcarpeta_path = os.path.join(FOTOS_VEHICULOS_DIR, subcarpeta_nombre)

                    # Crear la subcarpeta si no existe (aunque debería existir desde la salida)
                    if not os.path.exists(subcarpeta_path):
                        os.makedirs(subcarpeta_path)
                        logging.info(f"Subcarpeta creada: {subcarpeta_path}")

                    # Guardar las fotos de llegada en la subcarpeta
                    for i in range(1, 5):
                        foto_key = f"foto_km_final_{i}"
                        foto_fin = files.get(foto_key)
                        if foto_fin:
                            original_extension = os.path.splitext(foto_fin.filename)[1] if foto_fin.filename else ".jpg"
                            filename_fin = f"{subcarpeta_nombre}_llegada_{i}{original_extension}"
                            path_fin = os.path.join(subcarpeta_path, filename_fin)
                            foto_fin.save(path_fin)
                            logging.info(f"Foto de fin {i} guardada en {path_fin} para fila {ultimo_registro}")

                    return True, "Datos de llegada actualizados correctamente.", ultimo_registro
                else:
                    return False, "El último registro ya tiene datos de llegada. No puedes actualizarlo."
            else:
                return False, "No has enviado el Formulario de Datos de Salida correspondiente."

    except Exception as e:
        logging.error(f"Error al procesar datos de choferes: {str(e)}")
        return False, f"Error al procesar datos: {str(e)}"

@app.route('/api/recibir_datos_choferes', methods=['POST'])
def recibir_datos_choferes():
    """Recibe datos o fotos del formulario de choferes."""
    result = procesar_datos_choferes(request.form, request.files)
    if len(result) == 3:  # Caso con row_idx
        success, message, row_idx = result
        if success:
            return jsonify({"status": "success", "message": message, "row_idx": row_idx}), 200
        else:
            return jsonify({"status": "error", "message": message}), 400
    else:  # Caso sin row_idx
        success, message = result
        if success:
            return jsonify({"status": "success", "message": message}), 200
        else:
            return jsonify({"status": "error", "message": message}), 400


@app.route('/api/conductores', methods=['GET'])
def get_conductores():
    """Obtiene la lista de conductores."""
    try:
        if not os.path.exists(CONDUCTORES_CSV_PATH):
            return jsonify({"error": f"No se encontró el archivo de conductores"}), 404

        df = pd.read_csv(CONDUCTORES_CSV_PATH)
        if 'conductor' in df.columns:
            conductores = df['conductor'].dropna().astype(str).tolist()
        else:
            conductores = []
            logging.warning(f"La columna 'conductor' no se encontró en {CONDUCTORES_CSV_PATH}")
        return jsonify(conductores)

    except Exception as e:
        logging.error(f"Error al leer el archivo de conductores: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/vehiculos_info', methods=['GET'])
def get_vehiculos_info():
    """Obtiene la información de los vehículos (tipo y placa)."""
    try:
        if not os.path.exists(VEHICULOS_INFO_CSV_PATH):
            return jsonify({"error": f"No se encontró el archivo de vehículos"}), 404
        df = pd.read_csv(VEHICULOS_INFO_CSV_PATH)
        if 'tipo_vehiculo' in df.columns and 'placa' in df.columns:
            vehiculos = df[['tipo_vehiculo', 'placa']].dropna().astype(str).to_dict('records')
            return jsonify(vehiculos)
        else:
            missing_cols = []
            if 'tipo_vehiculo' not in df.columns:
                missing_cols.append('tipo_vehiculo')
            if 'placa' not in df.columns:
                missing_cols.append('placa')
            logging.warning(f"Faltan columnas en {VEHICULOS_INFO_CSV_PATH}: {', '.join(missing_cols)}")
            return jsonify({"error": f"Faltan columnas: {', '.join(missing_cols)}"}), 400

    except Exception as e:
        logging.error(f"Error al leer el archivo de vehículos: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/descargar-registro-rutas', methods=['GET'])
def descargar_registro_rutas():
    """Descarga el archivo Excel de registros de choferes."""
    try:
        logging.info(f"Intentando enviar archivo de registro de rutas: {REGISTROS_CHOFERES_EXCEL}")
        return send_file(REGISTROS_CHOFERES_EXCEL, as_attachment=True, download_name='registros_choferes.xlsx')
    except Exception as e:
        logging.error(f"Error al generar descarga de Excel de registros de choferes: {str(e)}")
        return str(e), 500

@app.route('/api/listar-carpetas-fotos', methods=['GET'])
def listar_carpetas_fotos():
    """Devuelve la lista de carpetas con el número de fotos en cada una."""
    try:
        carpetas = {}
        for nombre in os.listdir(FOTOS_VEHICULOS_DIR):
            carpeta_path = os.path.join(FOTOS_VEHICULOS_DIR, nombre)
            if os.path.isdir(carpeta_path):
                num_fotos = len([f for f in os.listdir(carpeta_path) if os.path.isfile(os.path.join(carpeta_path, f))])
                carpetas[nombre] = num_fotos
        return jsonify(carpetas)
    except Exception as e:
        logging.error(f"Error al listar carpetas de fotos: {str(e)}")
        return jsonify({"error": str(e)}), 500

@app.route('/descargar-carpeta-fotos/<nombre_carpeta>', methods=['GET'])
def descargar_carpeta_fotos_especifica(nombre_carpeta):
    """Comprime y descarga una carpeta específica de fotos_vehiculos."""
    try:
        carpeta_path = os.path.join(FOTOS_VEHICULOS_DIR, nombre_carpeta)
        if not os.path.exists(carpeta_path) or not os.path.isdir(carpeta_path):
            return jsonify({"error": "Carpeta no encontrada"}), 404
        
        zip_file_path = os.path.join(BASE_DIR, f"{nombre_carpeta}.zip")
        with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(carpeta_path):
                for file in files:
                    zipf.write(os.path.join(root, file),
                              os.path.relpath(os.path.join(root, file),
                                             os.path.join(carpeta_path, '..')))
        
        # Enviar el archivo ZIP al cliente
        response = send_file(zip_file_path, as_attachment=True, download_name=f"{nombre_carpeta}.zip")
        
        # Eliminar el archivo ZIP después de enviarlo
        os.remove(zip_file_path)
        logging.info(f"Archivo ZIP temporal eliminado: {zip_file_path}")
        
        return response
    except Exception as e:
        logging.error(f"Error al comprimir o descargar la carpeta {nombre_carpeta}: {str(e)}")
        return str(e), 500

@app.route('/descargar-carpeta-fotos', methods=['GET'])
def descargar_carpeta_fotos():
    """Comprime y descarga la carpeta de fotos de kilometraje."""
    try:
        zip_file_path = os.path.join(BASE_DIR, "fotos_vehiculos.zip")
        with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for root, _, files in os.walk(FOTOS_VEHICULOS_DIR):
                for file in files:
                    zipf.write(os.path.join(root, file),
                               os.path.relpath(os.path.join(root, file),
                                               os.path.join(FOTOS_VEHICULOS_DIR, '..')))
        logging.info(f"Intentando enviar carpeta de fotos comprimida: {zip_file_path}")
        return send_file(zip_file_path, as_attachment=True, download_name='fotos_vehiculos.zip')
    except Exception as e:
        logging.error(f"Error al comprimir o descargar la carpeta de fotos: {str(e)}")
        return str(e), 500


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=False) # debug=True para desarrollo